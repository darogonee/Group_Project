from openpyxl import load_workbook
import json, pyperclip

workbook = load_workbook(filename="exercises.xlsx")
sheet = workbook.active
exercises = []
for value in sheet.iter_rows(min_row=3,min_col=1, max_col=9, max_row=272,values_only=True):
    print(value)
    exercise_id = value[0]
    muscle_group = value[1]
    exercise_name = value[2]
    level = value[3]
    ulc = value[4]
    pp = value[5]
    modality = value[6]
    joint = value[7]
    equipment = value[8]

    exercises.append({"exercise_id":exercise_id, "muscle_group":muscle_group, "exercise_name":exercise_name, "level":level, "ulc":ulc, "pp":pp, "modality":modality, "joint":joint, "equipment":equipment})

pyperclip.copy(str(exercises))

# {exercise_id: {"muscle_group":muscle_group, "exercise":exercise, "level"}}

